from __future__ import annotations
from custom_auth.models import GoogleAssociation
from custom_auth.management.commands.linkgsuite import link_gsuite_users
from openpyxl.cell import cell
from openpyxl import styles
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from django.core.exceptions import ObjectDoesNotExist
import openpyxl
from openpyxl.utils import get_column_letter

from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from typing import Any, List, Iterator, Optional, Callable
    from django.contrib.admin import ModelAdmin
    from django.http import HttpRequest
    from django.db.models import QuerySet
    ExcelCellType = Any

def get_direct_prop(obj: Any, fields: List[str]) -> Any:
    """ Gets a model property of an object """

    # shouldn't happen, but whatever
    if len(fields) == 0:
        return obj

    # if we have a single field to get
    elif len(fields) == 1:
        field = fields[0]

        # we may have a display getter
        try:
            display_getter = 'get_{}_display'.format(field)
            if hasattr(obj, display_getter):
                return getattr(obj, display_getter)()
        except:
            pass

        # if we do not, return the field itself
        return getattr(obj, field)

    # else go recursively
    else:
        return get_direct_prop(getattr(obj, fields[0]), fields[1:])


def get_model_prop(modeladmin: ModelAdmin, obj: Any, field: str, default: Any = None) -> Any:
    """ Gets a model property or None"""
    try:
        try:
            return get_direct_prop(obj, field.split("__"))
        except AttributeError:
            try:
                attr = getattr(modeladmin, field)
                return attr(obj)
            except AttributeError:
                return default
    except ObjectDoesNotExist:
        return default


def to_excel(value: Any) -> ExcelCellType:
    """ Turns any value into a value understood by excel """

    # if we know the type, return it immediately
    if isinstance(value, cell.KNOWN_TYPES):
        return value

    # if we are none of the above and have a name, return the name
    elif hasattr(value, "name"):
        return to_excel(getattr(value, "name"))

    # if we are a list, return the list
    elif isinstance(value, list):
        return ', '.join(map(to_excel, value))

    # fallback to string
    else:
        return str(value)


def export_as_xslx_action(description: str = "Export selected objects as XSLX file",
                          fields: Optional[Iterator[str]] = None, header: bool = True) -> Callable[[ModelAdmin, HttpRequest, QuerySet], HttpResponse]:
    """
    Return an action that exports the given fields as XSLX files
    """

    def export_as_xslx(modeladmin: ModelAdmin, request: HttpRequest, queryset: QuerySet) -> HttpResponse:

        # get fields to export
        opts = modeladmin.model._meta
        if not fields:
            field_names = [field.name for field in opts.fields]
        else:
            field_names = fields

        # Create a response header
        response = HttpResponse(
            content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(
            str(opts).replace('.', '_'))

        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = str(opts).replace('.', '_')

        # Write the header (if desired)
        if header:
            def makeHeaderCell(field):
                c = cell.Cell(ws, value=field)
                c.font = styles.Font(bold=True)
                return c
            ws.append([makeHeaderCell(field) for field in field_names])

        # Write each of the rows
        for row in queryset.values_list(*field_names):
            def makeCell(prop):
                try:
                    return to_excel(prop)
                except:
                    return str(prop)
            ws.append([makeCell(c) for c in row])

        # adjust column widths
        # adapted from https://stackoverflow.com/a/39530676
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for c in col:
                try:
                    if len(str(c.value)) > max_length:
                        max_length = len(c.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width

        # and export
        wb.save(response)
        return response

    export_as_xslx.short_description = description
    return export_as_xslx


def link_to_gsuite_action(modeladmin: ModelAdmin, request: HttpRequest, queryset: QuerySet) -> None:
    """ Link to GSuite links users to a GSuite Account """
    profiles = get_user_model().objects.filter(alumni__in=queryset)

    link_gsuite_users(
        profiles, False, on_message=lambda x: modeladmin.message_user(request, x))


link_to_gsuite_action.short_description = 'Link Users to GSuite'


def unlink_from_gsuite_action(modeladmin: ModelAdmin, request: HttpRequest, queryset: QuerySet) -> None:
    """ Unlink Users from GSuite """

    count, _ = GoogleAssociation.objects.filter(
        user__alumni__in=queryset).delete()
    modeladmin.message_user(
        request, 'Unlinked {} user(s) from their GSuite Account(s). '.format(count))


unlink_from_gsuite_action.short_description = 'Unlink Users from GSuite'


class AlumniAdminActions:
    """ Actions available in the Django Alumni Admin Page """

    full_export_fields = (
        # Profile data
        'profile__username', 'profile__is_staff', 'profile__is_superuser',
        'profile__date_joined', 'profile__last_login',

        # Alumni Model
        'givenName', 'middleName', 'familyName', 'email', 'existingEmail',
        'resetExistingEmailPassword', 'sex', 'birthday',
        'nationality', 'category',

        # Address Data
        'address__address_line_1', 'address__address_line_2', 'address__city',
        'address__zip', 'address__state', 'address__country',

        # 'Social' Data
        'social__facebook', 'social__linkedin', 'social__twitter',
        'social__instagram', 'social__homepage',

        # 'Jacobs Data'
        'jacobs__college', 'jacobs__graduation', 'jacobs__degree',
        'jacobs__major', 'jacobs__comments',

        # 'Approval' Data
        'approval__approval', 'approval__time', 'approval__gsuite',

        # Job Data
        'job__employer', 'job__position', 'job__industry', 'job__job',

        # Skills Data
        'skills__otherDegrees', 'skills__spokenLanguages',
        'skills__programmingLanguages', 'skills__areasOfInterest',
        'skills__alumniMentor',

        # Membership Data
        'membership__tier', 'membership__desired_tier',

        # Atlas Settings
        'atlas__secret', 'atlas__included', 'atlas__birthdayVisible', 'atlas__contactInfoVisible',

        # Setup Data
        'setup__date',
    )

    actions = [
        export_as_xslx_action("Export as XSLX", fields=full_export_fields),
        link_to_gsuite_action,
        unlink_from_gsuite_action
    ]
